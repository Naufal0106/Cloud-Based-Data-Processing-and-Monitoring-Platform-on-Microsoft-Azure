import base64
import csv
import hashlib
import hmac
import io
import json
import logging
import math
import os
import re
import secrets
import time
import uuid
from collections import Counter
from datetime import datetime, timedelta, timezone
from email.parser import BytesParser
from email.policy import default as email_policy
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import azure.functions as func
from azure.cosmos import CosmosClient
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient

app = func.FunctionApp()

ALLOWED_STATUS = {"processed", "anomaly", "error"}
DEFAULT_LIMIT = 50
MAX_LIMIT = 200
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_UPLOAD_RECORDS = 1000
MAX_ANALYTICS_RECORDS = 500
MAX_PROFILE_COLUMNS = 40
MAX_CHART_BUCKETS = 8
SUPPORTED_UPLOAD_EXTENSIONS = {".json", ".csv", ".xlsx", ".xls"}
USER_ROLES = {"admin", "dev", "user"}
ADMIN_ROLE = "admin"
DEV_ROLE = "dev"
DEFAULT_USER_ROLE = "user"
AUTH_TOKEN_TTL_SECONDS = 60 * 60 * 8
PASSWORD_MIN_LENGTH = 8
PBKDF2_ITERATIONS = 160_000
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TELEMETRY_FILTER = "(NOT IS_DEFINED(c.doc_type) OR c.doc_type = 'telemetry')"

_cosmos_container = None
_users_container = None


def json_response(payload: dict[str, Any], status_code: int = 200) -> func.HttpResponse:
    return func.HttpResponse(
        json.dumps(payload, ensure_ascii=False),
        mimetype="application/json",
        status_code=status_code,
    )


def error_response(message: str, status_code: int = 500) -> func.HttpResponse:
    return json_response({"success": False, "error": message}, status_code)


def get_cosmos_container():
    global _cosmos_container

    if _cosmos_container is not None:
        return _cosmos_container

    try:
        key_vault_url = _required_env("KEY_VAULT_URL")
        db_name = _required_env("COSMOS_DATABASE")
        container_name = _required_env("COSMOS_CONTAINER")

        credential = DefaultAzureCredential()
        secret_client = SecretClient(vault_url=key_vault_url, credential=credential)
        cosmos_conn_str = secret_client.get_secret("cosmos-connection-string").value

        client = CosmosClient.from_connection_string(cosmos_conn_str)
        _cosmos_container = client.get_database_client(db_name).get_container_client(
            container_name
        )
        return _cosmos_container
    except Exception as exc:
        logging.exception("[Critical] Gagal inisialisasi koneksi Cosmos DB")
        raise exc


def get_users_container():
    global _users_container

    if _users_container is not None:
        return _users_container

    try:
        key_vault_url = _required_env("KEY_VAULT_URL")
        db_name = _required_env("COSMOS_DATABASE")
        container_name = _required_env("COSMOS_USERS_CONTAINER")

        credential = DefaultAzureCredential()
        secret_client = SecretClient(vault_url=key_vault_url, credential=credential)
        cosmos_conn_str = secret_client.get_secret("cosmos-connection-string").value

        client = CosmosClient.from_connection_string(cosmos_conn_str)
        _users_container = client.get_database_client(db_name).get_container_client(
            container_name
        )
        return _users_container
    except Exception as exc:
        logging.exception("[Critical] Gagal inisialisasi koneksi Cosmos DB users")
        raise exc


@app.blob_trigger(
    arg_name="myblob",
    path="raw-data/{name}",
    connection="AzureWebJobsStorage",
)
def process_blob(myblob: func.InputStream):
    blob_name = myblob.name
    logging.info("[BlobTrigger] File masuk: %s", blob_name)

    try:
        data = parse_file_payload(myblob.read(), blob_name)
    except ValueError as exc:
        logging.exception("[BlobTrigger] File %s tidak valid: %s", blob_name, exc)
        return

    try:
        processed_list = process_data(data, source_file=blob_name)
        attach_system_owner(processed_list)
        saved_count = save_to_cosmos(processed_list)
        logging.info("[BlobTrigger] %s selesai. %s record disimpan.", blob_name, saved_count)
    except Exception:
        logging.exception("[BlobTrigger] Error memproses %s", blob_name)


@app.route(route="data", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_data(req: func.HttpRequest) -> func.HttpResponse:
    claims, auth_error = get_auth_claims(req)
    if auth_error:
        return auth_error

    try:
        limit = parse_limit(req.params.get("limit"))
        status_filter = parse_status(req.params.get("status"))
    except ValueError as exc:
        return error_response(str(exc), 400)

    try:
        container = get_cosmos_container()
        query = f"SELECT TOP {limit} * FROM c WHERE {telemetry_scope_filter(claims)}"
        parameters = telemetry_scope_parameters(claims)

        if status_filter:
            query += " AND c.status = @status"
            parameters.append({"name": "@status", "value": status_filter})

        query += " ORDER BY c.processed_at DESC"

        items = list(
            container.query_items(
                query=query,
                parameters=parameters,
                enable_cross_partition_query=True,
            )
        )
        return json_response({"success": True, "count": len(items), "data": items})
    except Exception:
        logging.exception("[GET /data] Error")
        return error_response("Gagal mengambil data dari Cosmos DB")


@app.route(route="stats", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_stats(req: func.HttpRequest) -> func.HttpResponse:
    claims, auth_error = get_auth_claims(req)
    if auth_error:
        return auth_error

    try:
        container = get_cosmos_container()
        scope_filter = telemetry_scope_filter(claims)
        scope_params = telemetry_scope_parameters(claims)
        stats = {
            "total_records": count_items(container, scope_filter, scope_params),
            "processed": count_items(
                container, f"{scope_filter} AND c.status = @processed_status",
                scope_params + [{"name": "@processed_status", "value": "processed"}],
            ),
            "anomaly": count_items(
                container, f"{scope_filter} AND c.status = @anomaly_status",
                scope_params + [{"name": "@anomaly_status", "value": "anomaly"}],
            ),
            "errors": count_items(
                container, f"{scope_filter} AND c.status = @error_status",
                scope_params + [{"name": "@error_status", "value": "error"}],
            ),
            "categories": {
                "sensor": count_items(
                    container, f"{scope_filter} AND c.category = @sensor_category",
                    scope_params + [{"name": "@sensor_category", "value": "sensor"}],
                ),
                "log": count_items(
                    container, f"{scope_filter} AND c.category = @log_category",
                    scope_params + [{"name": "@log_category", "value": "log"}],
                ),
                "generic": count_items(
                    container, f"{scope_filter} AND c.category = @generic_category",
                    scope_params + [{"name": "@generic_category", "value": "generic"}],
                ),
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        return json_response({"success": True, "stats": stats})
    except Exception:
        logging.exception("[GET /stats] Error")
        return error_response("Gagal mengambil statistik dari Cosmos DB")


@app.route(route="upload", methods=["POST"], auth_level=func.AuthLevel.FUNCTION)
def upload_data(req: func.HttpRequest) -> func.HttpResponse:
    claims, auth_error = get_auth_claims(req)
    if auth_error:
        return auth_error

    try:
        data, source_file = parse_upload_request(req)
    except ValueError as exc:
        logging.warning("[POST /upload] Payload tidak valid: %s", exc)
        return error_response(str(exc), 400)

    if not is_valid_payload(data):
        return error_response("Payload harus berupa object atau array record", 400)

    try:
        clean_requested = parse_bool_param(req.params.get("clean"))
        cleaning = None
        if clean_requested:
            data, cleaning = clean_data_payload(data)

        processed = process_data(data, source_file=source_file)
        attach_owner(processed, claims)
        saved_count = save_to_cosmos(processed)
        analysis = build_data_science_payload(data, source_file, processed)
        return json_response(
            {
                "success": True,
                "message": f"{saved_count} record berhasil diproses dan disimpan.",
                "count": saved_count,
                "cleaned": clean_requested,
                "cleaning": cleaning,
                "profile": analysis["profile"],
                "quality": analysis["quality"],
                "charts": analysis["charts"],
                "recommendations": analysis["recommendations"],
            },
            201,
        )
    except Exception:
        logging.exception("[POST /upload] Error")
        return error_response("Gagal memproses dan menyimpan data")


@app.route(route="analyze", methods=["POST"], auth_level=func.AuthLevel.FUNCTION)
def analyze_upload(req: func.HttpRequest) -> func.HttpResponse:
    auth_error = require_auth(req)
    if auth_error:
        return auth_error

    try:
        data, source_file = parse_upload_request(req)
    except ValueError as exc:
        logging.warning("[POST /analyze] Payload tidak valid: %s", exc)
        return error_response(str(exc), 400)

    if not is_valid_payload(data):
        return error_response("Payload harus berupa object atau array record", 400)

    try:
        processed_preview = process_data(data, source_file=source_file)
        analysis = build_data_science_payload(data, source_file, processed_preview)
        return json_response({"success": True, **analysis})
    except Exception:
        logging.exception("[POST /analyze] Error")
        return error_response("Gagal menganalisis data")


@app.route(route="analytics", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_analytics(req: func.HttpRequest) -> func.HttpResponse:
    claims, auth_error = get_auth_claims(req)
    if auth_error:
        return auth_error

    try:
        limit = parse_analytics_limit(req.params.get("limit"))
    except ValueError as exc:
        return error_response(str(exc), 400)

    try:
        container = get_cosmos_container()
        parameters = telemetry_scope_parameters(claims)
        items = list(
            container.query_items(
                query=(
                    f"SELECT TOP {limit} * FROM c WHERE {telemetry_scope_filter(claims)} "
                    "ORDER BY c.processed_at DESC"
                ),
                parameters=parameters,
                enable_cross_partition_query=True,
            )
        )
        raw_records = [item.get("raw", item) if isinstance(item, dict) else item for item in items]
        analysis = build_data_science_payload(raw_records, "cosmos-telemetry", items)
        return json_response({"success": True, "count": len(items), **analysis})
    except Exception:
        logging.exception("[GET /analytics] Error")
        return error_response("Gagal mengambil analitik data")


@app.route(route="register", methods=["POST"], auth_level=func.AuthLevel.FUNCTION)
def register_user(req: func.HttpRequest) -> func.HttpResponse:
    try:
        payload = req.get_json()
    except ValueError:
        return error_response("Body harus berformat JSON valid", 400)

    try:
        name = validate_name(payload.get("name"))
        email = validate_email(payload.get("email"))
        password = validate_password(payload.get("password"))
    except (AttributeError, ValueError) as exc:
        return error_response(str(exc), 400)

    try:
        container = get_users_container()
        if find_user_by_email(container, email):
            return error_response("Email sudah terdaftar", 409)

        now = datetime.now(timezone.utc).isoformat()
        user = {
            "id": str(uuid.uuid4()),
            "doc_type": "user",
            "name": name,
            "email": email,
            "role": DEFAULT_USER_ROLE,
            "password_hash": hash_password(password),
            "created_at": now,
            "updated_at": now,
        }
        container.upsert_item(user)

        token = create_auth_token(user)
        return json_response(
            {
                "success": True,
                "message": "Registrasi berhasil",
                "token": token,
                "user": public_user(user),
            },
            201,
        )
    except Exception:
        logging.exception("[POST /register] Error")
        return error_response("Gagal melakukan registrasi")


@app.route(route="login", methods=["POST"], auth_level=func.AuthLevel.FUNCTION)
def login_user(req: func.HttpRequest) -> func.HttpResponse:
    try:
        payload = req.get_json()
    except ValueError:
        return error_response("Body harus berformat JSON valid", 400)

    try:
        email = validate_email(payload.get("email"))
        password = str(payload.get("password", ""))
    except (AttributeError, ValueError):
        return error_response("Email atau password salah", 401)

    try:
        container = get_users_container()
        user = find_user_by_email(container, email)
        if not user or not verify_password(password, user.get("password_hash", "")):
            return error_response("Email atau password salah", 401)

        user["last_login_at"] = datetime.now(timezone.utc).isoformat()
        container.upsert_item(user)

        token = create_auth_token(user)
        return json_response(
            {
                "success": True,
                "message": "Login berhasil",
                "token": token,
                "user": public_user(user),
            }
        )
    except Exception:
        logging.exception("[POST /login] Error")
        return error_response("Gagal melakukan login")


@app.route(route="me", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_current_user(req: func.HttpRequest) -> func.HttpResponse:
    claims, auth_error = get_auth_claims(req)
    if auth_error:
        return auth_error

    return json_response(
        {
            "success": True,
            "user": {
                "id": claims["sub"],
                "name": claims["name"],
                "email": claims["email"],
                "role": claims.get("role", "user"),
            },
        }
    )


@app.route(route="management/users", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def list_admin_users(req: func.HttpRequest) -> func.HttpResponse:
    _, auth_error = require_role(req, {ADMIN_ROLE})
    if auth_error:
        return auth_error

    try:
        container = get_users_container()
        users = list(
            container.query_items(
                query=(
                    "SELECT c.id, c.name, c.email, c.role, c.created_at, "
                    "c.updated_at, c.last_login_at FROM c "
                    "WHERE c.doc_type = 'user'"
                ),
                enable_cross_partition_query=True,
            )
        )
        users.sort(key=lambda item: item.get("created_at", ""), reverse=True)

        return json_response(
            {
                "success": True,
                "count": len(users),
                "users": [public_user(item, include_meta=True) for item in users],
            }
        )
    except Exception:
        logging.exception("[GET /management/users] Error")
        return error_response("Gagal mengambil daftar user")


@app.route(route="management/summary", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_management_summary(req: func.HttpRequest) -> func.HttpResponse:
    _, auth_error = require_role(req, {ADMIN_ROLE})
    if auth_error:
        return auth_error

    try:
        users_container = get_users_container()
        data_container = get_cosmos_container()
        recent_cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

        summary = {
            "users": {
                "total": count_items(users_container, "c.doc_type = 'user'"),
                "admin": count_items(users_container, "c.doc_type = 'user' AND c.role = @role", [{"name": "@role", "value": ADMIN_ROLE}]),
                "dev": count_items(users_container, "c.doc_type = 'user' AND c.role = @role", [{"name": "@role", "value": DEV_ROLE}]),
                "regular": count_items(users_container, "c.doc_type = 'user' AND c.role = @role", [{"name": "@role", "value": DEFAULT_USER_ROLE}]),
                "recent_login": count_items(
                    users_container,
                    "c.doc_type = 'user' AND IS_DEFINED(c.last_login_at) AND c.last_login_at >= @cutoff",
                    [{"name": "@cutoff", "value": recent_cutoff}],
                ),
            },
            "telemetry": {
                "total": count_items(data_container, TELEMETRY_FILTER),
                "processed": count_items(data_container, f"{TELEMETRY_FILTER} AND c.status = @status", [{"name": "@status", "value": "processed"}]),
                "anomaly": count_items(data_container, f"{TELEMETRY_FILTER} AND c.status = @status", [{"name": "@status", "value": "anomaly"}]),
                "error": count_items(data_container, f"{TELEMETRY_FILTER} AND c.status = @status", [{"name": "@status", "value": "error"}]),
            },
            "controls": {
                "public_register_role": DEFAULT_USER_ROLE,
                "admin_role_protected": True,
                "user_data_scoped": True,
                "management_route": "/api/management",
            },
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        return json_response({"success": True, "summary": summary})
    except Exception:
        logging.exception("[GET /management/summary] Error")
        return error_response("Gagal mengambil ringkasan admin")


@app.route(
    route="management/users/{user_id}/role",
    methods=["PATCH", "POST"],
    auth_level=func.AuthLevel.FUNCTION,
)
def update_admin_user_role(req: func.HttpRequest) -> func.HttpResponse:
    claims, auth_error = require_role(req, {ADMIN_ROLE})
    if auth_error:
        return auth_error

    user_id = str(req.route_params.get("user_id", "")).strip()
    if not user_id:
        return error_response("User id wajib diisi", 400)

    try:
        payload = req.get_json()
        role = validate_role(payload.get("role"))
    except (AttributeError, ValueError) as exc:
        return error_response(str(exc), 400)

    try:
        container = get_users_container()
        user = find_user_by_id(container, user_id)
        if not user:
            return error_response("User tidak ditemukan", 404)

        if user["id"] == claims["sub"] and role != ADMIN_ROLE:
            return error_response("Admin tidak bisa menurunkan role akunnya sendiri", 400)

        user["role"] = role
        user["updated_at"] = datetime.now(timezone.utc).isoformat()
        container.upsert_item(user)

        return json_response(
            {
                "success": True,
                "message": "Role user berhasil diperbarui",
                "user": public_user(user, include_meta=True),
            }
        )
    except Exception:
        logging.exception("[PATCH /management/users/{user_id}/role] Error")
        return error_response("Gagal memperbarui role user")


@app.route(route="dev/ops-summary", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_dev_ops_summary(req: func.HttpRequest) -> func.HttpResponse:
    _, auth_error = require_role(req, {ADMIN_ROLE, DEV_ROLE})
    if auth_error:
        return auth_error

    return build_ops_summary_response()


@app.route(route="management/ops-summary", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
def get_admin_ops_summary(req: func.HttpRequest) -> func.HttpResponse:
    _, auth_error = require_role(req, {ADMIN_ROLE, DEV_ROLE})
    if auth_error:
        return auth_error

    return build_ops_summary_response()


def build_ops_summary_response() -> func.HttpResponse:
    return json_response(
        {
            "success": True,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "azure": build_azure_ops_summary(),
            "cloudflare": build_cloudflare_ops_summary(),
        }
    )


@app.route(route="hello", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def hello(req: func.HttpRequest) -> func.HttpResponse:
    return json_response(
        {
            "success": True,
            "service": "K11 Monitoring Engine",
            "status": "online",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
    )


def parse_upload_request(req: func.HttpRequest) -> tuple[Any, str]:
    content_type = get_header(req, "content-type").lower()

    if "multipart/form-data" in content_type:
        filename, file_bytes, file_content_type = extract_multipart_file(req)
        return parse_file_payload(file_bytes, filename, file_content_type), filename

    if "application/json" in content_type or not content_type:
        try:
            payload = req.get_json()
        except ValueError as exc:
            raise ValueError("Body JSON tidak valid") from exc
        if not is_valid_payload(payload):
            raise ValueError("JSON harus berupa object atau array")
        return validate_json_records(payload), "http-upload.json"

    filename = req.params.get("filename") or guess_filename_from_content_type(content_type)
    return parse_file_payload(req.get_body(), filename, content_type), filename


def extract_multipart_file(req: func.HttpRequest) -> tuple[str, bytes, str]:
    content_type = get_header(req, "content-type")
    body = req.get_body()
    if len(body) > MAX_UPLOAD_BYTES:
        raise ValueError(f"Ukuran file maksimal {format_size(MAX_UPLOAD_BYTES)}")

    message = BytesParser(policy=email_policy).parsebytes(
        (
            f"Content-Type: {content_type}\r\n"
            "MIME-Version: 1.0\r\n\r\n"
        ).encode()
        + body
    )

    for part in message.iter_parts():
        filename = part.get_filename()
        field_name = part.get_param("name", header="content-disposition")
        if filename and field_name in {"file", "upload", "data"}:
            file_bytes = part.get_payload(decode=True) or b""
            file_type = part.get_content_type()
            if not file_bytes:
                raise ValueError("File upload kosong")
            return sanitize_filename(filename), file_bytes, file_type

    for part in message.iter_parts():
        filename = part.get_filename()
        if filename:
            file_bytes = part.get_payload(decode=True) or b""
            if not file_bytes:
                raise ValueError("File upload kosong")
            return sanitize_filename(filename), file_bytes, part.get_content_type()

    raise ValueError("Field file tidak ditemukan pada multipart upload")


def parse_file_payload(
    file_bytes: bytes,
    filename: str,
    content_type: str | None = None,
) -> Any:
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise ValueError(f"Ukuran file maksimal {format_size(MAX_UPLOAD_BYTES)}")

    extension = get_file_extension(filename, content_type)
    if extension not in SUPPORTED_UPLOAD_EXTENSIONS:
        allowed = ", ".join(sorted(SUPPORTED_UPLOAD_EXTENSIONS))
        raise ValueError(f"Format file tidak didukung. Gunakan {allowed}")

    if extension == ".json":
        return parse_json_file(file_bytes)
    if extension == ".csv":
        return parse_csv_file(file_bytes)
    if extension == ".xlsx":
        return parse_xlsx_file(file_bytes)
    if extension == ".xls":
        return parse_xls_file(file_bytes)

    raise ValueError("Format file tidak didukung")


def parse_json_file(file_bytes: bytes) -> Any:
    try:
        data = json.loads(decode_text_file(file_bytes))
    except json.JSONDecodeError as exc:
        raise ValueError("JSON tidak valid") from exc

    if not is_valid_payload(data):
        raise ValueError("JSON harus berupa object atau array")
    return validate_json_records(data)


def parse_csv_file(file_bytes: bytes) -> list[dict[str, Any]]:
    text = decode_text_file(file_bytes)
    sample = text[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = clean_headers(reader.fieldnames or [])
    if not headers:
        raise ValueError("CSV harus memiliki header pada baris pertama")

    records = []
    for row in reader:
        record = {}
        for raw_key, value in row.items():
            if raw_key is None:
                continue
            key = normalize_header(raw_key)
            if key:
                record[key] = normalize_cell(value)
        if has_record_value(record):
            records.append(record)

    return validate_tabular_records(records, "CSV")


def parse_xlsx_file(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Dependency openpyxl belum terpasang untuk membaca .xlsx") from exc

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("File Excel .xlsx tidak valid") from exc

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = clean_headers(next(rows))
    except StopIteration as exc:
        raise ValueError("File Excel kosong") from exc

    if not headers:
        raise ValueError("Excel harus memiliki header pada baris pertama")

    records = []
    for row in rows:
        record = row_to_record(headers, row)
        if has_record_value(record):
            records.append(record)

    return validate_tabular_records(records, "Excel")


def parse_xls_file(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Dependency xlrd belum terpasang untuk membaca .xls") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        worksheet = workbook.sheet_by_index(0)
    except Exception as exc:
        raise ValueError("File Excel .xls tidak valid") from exc

    if worksheet.nrows == 0:
        raise ValueError("File Excel kosong")

    headers = clean_headers(worksheet.row_values(0))
    if not headers:
        raise ValueError("Excel harus memiliki header pada baris pertama")

    records = []
    for row_index in range(1, worksheet.nrows):
        values = [
            normalize_xls_cell(
                worksheet.cell_value(row_index, col_index),
                worksheet.cell_type(row_index, col_index),
                workbook.datemode,
            )
            for col_index in range(worksheet.ncols)
        ]
        record = row_to_record(headers, values)
        if has_record_value(record):
            records.append(record)

    return validate_tabular_records(records, "Excel")


def row_to_record(headers: list[str], values: Any) -> dict[str, Any]:
    record = {}
    values_list = list(values or [])
    for index, key in enumerate(headers):
        if key:
            record[key] = normalize_cell(values_list[index] if index < len(values_list) else None)
    return record


def validate_tabular_records(records: list[dict[str, Any]], label: str) -> list[dict[str, Any]]:
    if not records:
        raise ValueError(f"{label} tidak memiliki baris data")
    if len(records) > MAX_UPLOAD_RECORDS:
        raise ValueError(f"{label} maksimal {MAX_UPLOAD_RECORDS} baris per upload")
    return records


def validate_json_records(data: Any) -> Any:
    record_count = len(data) if isinstance(data, list) else 1
    if record_count > MAX_UPLOAD_RECORDS:
        raise ValueError(f"JSON maksimal {MAX_UPLOAD_RECORDS} record per upload")
    return data


def clean_headers(raw_headers: Any) -> list[str]:
    headers = [normalize_header(header) for header in list(raw_headers or [])]
    return deduplicate_headers(headers)


def normalize_header(value: Any) -> str:
    header = str(value or "").strip().lower()
    header = re.sub(r"\s+", "_", header)
    header = re.sub(r"[^A-Za-z0-9_.-]", "_", header)
    header = re.sub(r"_+", "_", header).strip("_")
    return header


def deduplicate_headers(headers: list[str]) -> list[str]:
    seen = {}
    deduped = []

    for header in headers:
        if not header:
            deduped.append("")
            continue

        count = seen.get(header, 0)
        seen[header] = count + 1
        deduped.append(header if count == 0 else f"{header}_{count + 1}")

    return deduped


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def normalize_xls_cell(value: Any, cell_type: int, datemode: int) -> Any:
    try:
        import xlrd
    except ImportError:
        return normalize_cell(value)

    if cell_type == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(value, datemode).isoformat()
        except Exception:
            return normalize_cell(value)

    if cell_type == xlrd.XL_CELL_NUMBER and float(value).is_integer():
        return int(value)

    return normalize_cell(value)


def has_record_value(record: dict[str, Any]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in record.values())


def has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def decode_text_file(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("File teks tidak dapat dibaca")


def get_file_extension(filename: str, content_type: str | None = None) -> str:
    _, extension = os.path.splitext(filename.lower())
    if extension:
        return extension

    content_type = (content_type or "").split(";")[0].strip().lower()
    return {
        "application/json": ".json",
        "text/json": ".json",
        "text/csv": ".csv",
        "application/csv": ".csv",
        "application/vnd.ms-excel": ".xls",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    }.get(content_type, "")


def guess_filename_from_content_type(content_type: str) -> str:
    extension = get_file_extension("", content_type) or ".bin"
    return f"http-upload{extension}"


def sanitize_filename(filename: str) -> str:
    clean_name = os.path.basename(filename.replace("\\", "/")).strip()
    return clean_name or "upload.bin"


def get_header(req: func.HttpRequest, name: str) -> str:
    for key, value in req.headers.items():
        if key.lower() == name.lower():
            return value
    return ""


def format_size(bytes_count: int) -> str:
    if bytes_count < 1024 * 1024:
        return f"{bytes_count // 1024} KB"
    return f"{bytes_count // 1024 // 1024} MB"


def build_data_science_payload(
    data: Any,
    source_file: str,
    processed_records: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    records = to_record_list(data)
    profile = build_data_profile(records)
    quality = build_quality_report(records, profile)
    _, cleaning = clean_records(records, profile)

    return {
        "source_file": source_file,
        "profile": profile,
        "quality": quality,
        "cleaning": cleaning,
        "recommendations": build_cleaning_recommendations(quality, cleaning),
        "charts": build_chart_payload(records, processed_records or []),
        "sample": records[:5],
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


def to_record_list(data: Any) -> list[dict[str, Any]]:
    raw_records = data if isinstance(data, list) else [data]
    records = []

    for item in raw_records:
        if isinstance(item, dict):
            records.append(item)
        else:
            records.append({"value": item})

    return records


def build_data_profile(records: list[dict[str, Any]]) -> dict[str, Any]:
    columns = collect_columns(records)
    profiled_columns = [
        profile_column(column, records)
        for column in columns[:MAX_PROFILE_COLUMNS]
    ]

    missing_cells = sum(column["missing_count"] for column in profiled_columns)
    total_cells = len(records) * len(columns)
    numeric_columns = [
        column["name"]
        for column in profiled_columns
        if column["data_type"] == "numeric"
    ]
    categorical_columns = [
        column["name"]
        for column in profiled_columns
        if column["data_type"] in {"text", "boolean"}
    ]

    return {
        "record_count": len(records),
        "column_count": len(columns),
        "profiled_column_count": len(profiled_columns),
        "truncated": len(columns) > MAX_PROFILE_COLUMNS,
        "missing_cells": missing_cells,
        "missing_cell_pct": round((missing_cells / total_cells) * 100, 2)
        if total_cells
        else 0,
        "duplicate_rows": count_duplicate_records(records),
        "numeric_columns": numeric_columns,
        "categorical_columns": categorical_columns,
        "columns": profiled_columns,
    }


def collect_columns(records: list[dict[str, Any]]) -> list[str]:
    seen = set()
    columns = []

    for record in records:
        for key in record.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)

    return columns


def profile_column(column: str, records: list[dict[str, Any]]) -> dict[str, Any]:
    values = [record.get(column) for record in records]
    present_values = [value for value in values if has_value(value)]
    missing_count = len(values) - len(present_values)
    type_counts = Counter(detect_value_type(value) for value in present_values)
    numeric_values = [
        number
        for value in present_values
        if (number := coerce_number(value)) is not None
    ]
    datetime_values = [
        value
        for value in present_values
        if detect_value_type(value) == "datetime"
    ]
    boolean_values = [
        value
        for value in present_values
        if detect_value_type(value) == "boolean"
    ]

    data_type = infer_column_type(
        present_count=len(present_values),
        type_counts=type_counts,
        numeric_count=len(numeric_values),
        datetime_count=len(datetime_values),
        boolean_count=len(boolean_values),
    )
    top_values = Counter(canonical_preview(value) for value in present_values).most_common(5)
    profile = {
        "name": column,
        "data_type": data_type,
        "missing_count": missing_count,
        "missing_pct": round((missing_count / len(values)) * 100, 2) if values else 0,
        "unique_count": len({canonical_record(value) for value in present_values}),
        "top_values": [
            {"value": value, "count": count}
            for value, count in top_values
        ],
        "sample_values": [safe_preview(value) for value in present_values[:3]],
        "type_counts": dict(type_counts),
    }

    if numeric_values:
        profile["numeric"] = summarize_numbers(numeric_values)

    return profile


def infer_column_type(
    present_count: int,
    type_counts: Counter,
    numeric_count: int,
    datetime_count: int,
    boolean_count: int,
) -> str:
    if present_count == 0:
        return "empty"

    if numeric_count / present_count >= 0.8:
        return "numeric"
    if boolean_count / present_count >= 0.8:
        return "boolean"
    if datetime_count / present_count >= 0.8:
        return "datetime"
    if len(type_counts) > 1:
        return "mixed"

    return next(iter(type_counts), "text")


def summarize_numbers(values: list[float]) -> dict[str, float]:
    sorted_values = sorted(values)
    count = len(sorted_values)
    total = sum(sorted_values)
    average = total / count if count else 0
    variance = sum((value - average) ** 2 for value in sorted_values) / count if count else 0
    middle = count // 2
    median_value = (
        sorted_values[middle]
        if count % 2
        else (sorted_values[middle - 1] + sorted_values[middle]) / 2
    )

    return {
        "min": round(sorted_values[0], 4),
        "max": round(sorted_values[-1], 4),
        "mean": round(average, 4),
        "median": round(median_value, 4),
        "stddev": round(math.sqrt(variance), 4),
    }


def build_quality_report(
    records: list[dict[str, Any]],
    profile: dict[str, Any],
) -> dict[str, Any]:
    record_count = profile["record_count"]
    column_count = profile["column_count"]
    missing_cells = profile["missing_cells"]
    duplicate_rows = profile["duplicate_rows"]
    invalid_temperature_rows = count_invalid_temperature(records)
    mixed_columns = [
        column["name"]
        for column in profile["columns"]
        if column["data_type"] == "mixed"
    ]
    empty_columns = [
        column["name"]
        for column in profile["columns"]
        if column["data_type"] == "empty"
    ]
    issues = []

    if missing_cells:
        issues.append(
            {
                "type": "missing_values",
                "severity": "warn",
                "message": f"{missing_cells} cell kosong/null ditemukan.",
                "affected": missing_cells,
            }
        )
    if duplicate_rows:
        issues.append(
            {
                "type": "duplicate_rows",
                "severity": "warn",
                "message": f"{duplicate_rows} baris duplikat terdeteksi.",
                "affected": duplicate_rows,
            }
        )
    if invalid_temperature_rows:
        issues.append(
            {
                "type": "invalid_temperature",
                "severity": "error",
                "message": f"{invalid_temperature_rows} record memiliki temperature tidak numerik.",
                "affected": invalid_temperature_rows,
            }
        )
    if mixed_columns:
        issues.append(
            {
                "type": "mixed_types",
                "severity": "warn",
                "message": "Ada kolom dengan tipe data campuran.",
                "columns": mixed_columns[:10],
                "affected": len(mixed_columns),
            }
        )
    if empty_columns:
        issues.append(
            {
                "type": "empty_columns",
                "severity": "info",
                "message": "Ada kolom yang seluruh nilainya kosong.",
                "columns": empty_columns[:10],
                "affected": len(empty_columns),
            }
        )

    total_cells = max(1, record_count * max(1, column_count))
    missing_ratio = missing_cells / total_cells
    duplicate_ratio = duplicate_rows / max(1, record_count)
    invalid_ratio = invalid_temperature_rows / max(1, record_count)
    mixed_ratio = len(mixed_columns) / max(1, column_count)
    score = 100
    score -= min(35, missing_ratio * 70)
    score -= min(25, duplicate_ratio * 60)
    score -= min(25, invalid_ratio * 80)
    score -= min(15, mixed_ratio * 45)
    score = max(0, round(score))

    return {
        "score": score,
        "dirty": any(issue["severity"] in {"warn", "error"} for issue in issues),
        "issue_count": len(issues),
        "missing_cells": missing_cells,
        "duplicate_rows": duplicate_rows,
        "invalid_temperature_rows": invalid_temperature_rows,
        "mixed_columns": mixed_columns,
        "empty_columns": empty_columns,
        "issues": issues,
    }


def clean_data_payload(data: Any) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    records = to_record_list(data)
    profile = build_data_profile(records)
    cleaned_records, cleaning = clean_records(records, profile)
    if not cleaned_records:
        raise ValueError("Data kosong setelah proses cleaning")
    return cleaned_records, cleaning


def clean_records(
    records: list[dict[str, Any]],
    profile: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    profile_map = {column["name"]: column for column in profile["columns"]}
    numeric_columns = {
        column["name"]
        for column in profile["columns"]
        if column["data_type"] == "numeric" and not is_identifier_column(column["name"])
    }
    boolean_columns = {
        column["name"]
        for column in profile["columns"]
        if column["data_type"] == "boolean"
    }
    cleaning = {
        "records_before": len(records),
        "records_after": 0,
        "trimmed_values": 0,
        "blank_cells_to_null": 0,
        "numeric_conversions": 0,
        "boolean_conversions": 0,
        "empty_rows_removed": 0,
        "duplicates_removed": 0,
    }
    cleaned_records = []
    seen = set()

    for record in records:
        cleaned = {}
        for key, value in record.items():
            cleaned[key] = clean_value(
                key,
                value,
                numeric_columns,
                boolean_columns,
                profile_map,
                cleaning,
            )

        if not has_record_value(cleaned):
            cleaning["empty_rows_removed"] += 1
            continue

        canonical = canonical_record(cleaned)
        if canonical in seen:
            cleaning["duplicates_removed"] += 1
            continue

        seen.add(canonical)
        cleaned_records.append(cleaned)

    cleaning["records_after"] = len(cleaned_records)
    return cleaned_records, cleaning


def clean_value(
    key: str,
    value: Any,
    numeric_columns: set[str],
    boolean_columns: set[str],
    profile_map: dict[str, dict[str, Any]],
    cleaning: dict[str, int],
) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        if stripped != value:
            cleaning["trimmed_values"] += 1
        if stripped == "":
            cleaning["blank_cells_to_null"] += 1
            return None

        if key in numeric_columns:
            number = coerce_number(stripped)
            if number is not None:
                cleaning["numeric_conversions"] += int(not isinstance(value, (int, float)))
                return number

        if key in boolean_columns:
            boolean = coerce_boolean(stripped)
            if boolean is not None:
                cleaning["boolean_conversions"] += 1
                return boolean

        return stripped

    if isinstance(value, list):
        return [
            clean_value(key, item, numeric_columns, boolean_columns, profile_map, cleaning)
            for item in value
        ]

    return value


def build_cleaning_recommendations(
    quality: dict[str, Any],
    cleaning: dict[str, Any],
) -> list[str]:
    recommendations = []

    if quality["missing_cells"]:
        recommendations.append("Ubah cell kosong menjadi null agar query dan chart konsisten.")
    if quality["duplicate_rows"]:
        recommendations.append("Hapus baris duplikat sebelum data disimpan.")
    if quality["invalid_temperature_rows"]:
        recommendations.append("Perbaiki kolom temperature agar bernilai angka.")
    if quality["mixed_columns"]:
        recommendations.append("Standarkan tipe data pada kolom campuran.")
    if cleaning["numeric_conversions"]:
        recommendations.append("Konversi angka berbentuk teks menjadi numeric value.")

    return recommendations or ["Data siap diproses tanpa cleaning tambahan."]


def build_chart_payload(
    records: list[dict[str, Any]],
    processed_records: list[dict[str, Any]],
) -> dict[str, Any]:
    profile = build_data_profile(records)
    return {
        "status_distribution": counter_chart(
            Counter(
                str(record.get("status", "unknown"))
                for record in processed_records
                if isinstance(record, dict)
            )
        ),
        "category_distribution": counter_chart(
            Counter(
                str(record.get("category", "generic"))
                for record in processed_records
                if isinstance(record, dict)
            )
        ),
        "missing_by_column": missing_chart(profile),
        "numeric_histograms": numeric_histograms(profile, records),
        "top_values": categorical_top_values(profile),
        "correlation": correlation_matrix(profile, records),
    }


def counter_chart(counter: Counter) -> dict[str, list[Any]]:
    items = counter.most_common()
    return {
        "labels": [label for label, _ in items],
        "data": [count for _, count in items],
    }


def missing_chart(profile: dict[str, Any]) -> dict[str, list[Any]]:
    items = sorted(
        (
            (column["name"], column["missing_count"])
            for column in profile["columns"]
            if column["missing_count"] > 0
        ),
        key=lambda item: item[1],
        reverse=True,
    )[:12]

    return {
        "labels": [name for name, _ in items],
        "data": [count for _, count in items],
    }


def numeric_histograms(
    profile: dict[str, Any],
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    histograms = []

    for column in profile["columns"]:
        if column["data_type"] != "numeric":
            continue

        values = [
            number
            for record in records
            if (number := coerce_number(record.get(column["name"]))) is not None
        ]
        if values:
            histograms.append(build_histogram(column["name"], values))
        if len(histograms) >= 3:
            break

    return histograms


def build_histogram(column: str, values: list[float]) -> dict[str, Any]:
    min_value = min(values)
    max_value = max(values)

    if min_value == max_value:
        return {"column": column, "labels": [format_number(min_value)], "data": [len(values)]}

    bucket_count = min(MAX_CHART_BUCKETS, max(1, len(values)))
    step = (max_value - min_value) / bucket_count
    buckets = [0 for _ in range(bucket_count)]

    for value in values:
        index = min(bucket_count - 1, int((value - min_value) / step))
        buckets[index] += 1

    labels = []
    for index in range(bucket_count):
        start = min_value + (step * index)
        end = start + step
        labels.append(f"{format_number(start)}-{format_number(end)}")

    return {"column": column, "labels": labels, "data": buckets}


def categorical_top_values(profile: dict[str, Any]) -> list[dict[str, Any]]:
    charts = []
    for column in profile["columns"]:
        if column["data_type"] not in {"text", "boolean", "datetime"}:
            continue
        if not column["top_values"]:
            continue

        charts.append(
            {
                "column": column["name"],
                "labels": [item["value"] for item in column["top_values"]],
                "data": [item["count"] for item in column["top_values"]],
            }
        )
        if len(charts) >= 3:
            break

    return charts


def correlation_matrix(
    profile: dict[str, Any],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    numeric_columns = [
        column["name"]
        for column in profile["columns"]
        if column["data_type"] == "numeric"
    ][:6]
    matrix = []

    for left in numeric_columns:
        row = []
        for right in numeric_columns:
            row.append(pearson_for_columns(records, left, right))
        matrix.append(row)

    return {"labels": numeric_columns, "matrix": matrix}


def pearson_for_columns(
    records: list[dict[str, Any]],
    left: str,
    right: str,
) -> float | None:
    pairs = []
    for record in records:
        left_value = coerce_number(record.get(left))
        right_value = coerce_number(record.get(right))
        if left_value is not None and right_value is not None:
            pairs.append((left_value, right_value))

    if len(pairs) < 2:
        return None

    xs = [pair[0] for pair in pairs]
    ys = [pair[1] for pair in pairs]
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    numerator = sum((x - mean_x) * (y - mean_y) for x, y in pairs)
    denominator_x = math.sqrt(sum((x - mean_x) ** 2 for x in xs))
    denominator_y = math.sqrt(sum((y - mean_y) ** 2 for y in ys))
    if denominator_x == 0 or denominator_y == 0:
        return 1 if left == right else None
    return round(numerator / (denominator_x * denominator_y), 4)


def count_duplicate_records(records: list[dict[str, Any]]) -> int:
    seen = set()
    duplicates = 0

    for record in records:
        canonical = canonical_record(record)
        if canonical in seen:
            duplicates += 1
        else:
            seen.add(canonical)

    return duplicates


def count_invalid_temperature(records: list[dict[str, Any]]) -> int:
    invalid = 0
    for record in records:
        if not isinstance(record, dict):
            continue
        for key in ("temperature", "temp", "suhu"):
            value = record.get(key)
            if has_value(value) and coerce_number(value) is None:
                invalid += 1
                break
    return invalid


def detect_value_type(value: Any) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "numeric" if math.isfinite(float(value)) else "text"
    if isinstance(value, str):
        stripped = value.strip()
        if coerce_boolean(stripped) is not None:
            return "boolean"
        if coerce_number(stripped) is not None:
            return "numeric"
        if coerce_datetime(stripped):
            return "datetime"
        return "text"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, (dict, list)):
        return "object"
    return "text"


def coerce_number(value: Any) -> float | int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if not math.isfinite(number):
            return None
        return int(number) if number.is_integer() else number
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if re.fullmatch(r"-?\d+,\d+", text):
        text = text.replace(",", ".")
    if not re.fullmatch(r"-?\d+(\.\d+)?", text):
        return None

    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def coerce_boolean(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if not isinstance(value, str):
        return None

    lowered = value.strip().lower()
    if lowered in {"true", "yes", "y", "ya"}:
        return True
    if lowered in {"false", "no", "n", "tidak"}:
        return False
    return None


def coerce_datetime(value: str) -> bool:
    try:
        datetime.fromisoformat(value.replace("Z", "+00:00"))
        return True
    except ValueError:
        return False


def is_identifier_column(column: str) -> bool:
    lowered = column.lower()
    return any(token in lowered for token in ("id", "code", "kode", "phone", "tel", "zip"))


def canonical_record(value: Any) -> str:
    try:
        return json.dumps(value, sort_keys=True, ensure_ascii=False, default=str)
    except TypeError:
        return str(value)


def canonical_preview(value: Any) -> str:
    text = safe_preview(value)
    return text if len(text) <= 80 else f"{text[:77]}..."


def safe_preview(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, (dict, list)):
        return canonical_record(value)
    return str(value)


def format_number(value: float | int) -> str:
    return f"{value:g}"


def parse_bool_param(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "clean"}


def parse_analytics_limit(raw_limit: str | None) -> int:
    if raw_limit is None or raw_limit == "":
        return DEFAULT_LIMIT

    try:
        limit = int(raw_limit)
    except (TypeError, ValueError):
        raise ValueError("Parameter limit harus berupa angka")

    if limit < 1:
        raise ValueError("Parameter limit minimal 1")
    if limit > MAX_ANALYTICS_RECORDS:
        raise ValueError(f"Parameter limit maksimal {MAX_ANALYTICS_RECORDS}")
    return limit


def process_data(data: Any, source_file: str) -> list[dict[str, Any]]:
    records = data if isinstance(data, list) else [data]
    processed = []

    for item in records:
        record = build_base_record(item, source_file)

        if not isinstance(item, dict):
            record["category"] = "generic"
            record["summary"] = f"Record dari {source_file}"
            processed.append(record)
            continue

        if has_value(item.get("temperature")):
            enrich_sensor_record(record, item)
        elif has_value(item.get("level")) and has_value(item.get("message")):
            enrich_log_record(record, item)
        else:
            record["category"] = "generic"
            record["summary"] = f"Record dari {source_file}"

        processed.append(record)

    return processed


def telemetry_scope_filter(claims: dict[str, Any] | None) -> str:
    if claims and str(claims.get("role", "")).lower() == ADMIN_ROLE:
        return TELEMETRY_FILTER
    return f"{TELEMETRY_FILTER} AND c.owner_user_id = @owner_user_id"


def telemetry_scope_parameters(claims: dict[str, Any] | None) -> list[dict[str, Any]]:
    if claims and str(claims.get("role", "")).lower() == ADMIN_ROLE:
        return []
    return [{"name": "@owner_user_id", "value": str((claims or {}).get("sub", ""))}]


def attach_owner(records: list[dict[str, Any]], claims: dict[str, Any] | None) -> None:
    user_id = str((claims or {}).get("sub", ""))
    user_email = str((claims or {}).get("email", ""))
    user_role = str((claims or {}).get("role", DEFAULT_USER_ROLE)).lower()

    for record in records:
        record["owner_user_id"] = user_id
        record["owner_email"] = user_email
        record["owner_role"] = user_role if user_role in USER_ROLES else DEFAULT_USER_ROLE


def attach_system_owner(records: list[dict[str, Any]]) -> None:
    for record in records:
        record["owner_user_id"] = "system"
        record["owner_email"] = ""
        record["owner_role"] = "system"


def build_base_record(item: Any, source_file: str) -> dict[str, Any]:
    return {
        "id": str(uuid.uuid4()),
        "doc_type": "telemetry",
        "deviceId": extract_device_id(item),
        "source_file": source_file,
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "status": "processed",
        "category": "generic",
        "raw": item,
    }


def enrich_sensor_record(record: dict[str, Any], item: dict[str, Any]) -> None:
    record["category"] = "sensor"
    temperature = item.get("temperature")

    try:
        temperature_value = float(temperature)
    except (TypeError, ValueError):
        record["status"] = "error"
        record["summary"] = f"Nilai suhu tidak valid: {temperature}"
        record["alert"] = "Field temperature harus berupa angka"
        return

    record["summary"] = f"Suhu: {temperature_value:g} C"
    if temperature_value > 80:
        record["status"] = "anomaly"
        record["alert"] = "Suhu kritis, lebih dari 80 C"


def enrich_log_record(record: dict[str, Any], item: dict[str, Any]) -> None:
    record["category"] = "log"
    level = str(item.get("level", "")).upper()
    message = str(item.get("message", ""))

    record["summary"] = f"[{level}] {message}"
    if level in {"ERROR", "CRITICAL"}:
        record["status"] = "error"
        record["alert"] = "Log level kritis terdeteksi"


def save_to_cosmos(records: list[dict[str, Any]]) -> int:
    container = get_cosmos_container()

    for record in records:
        container.upsert_item(record)
        logging.info(
            "[Cosmos] Disimpan: id=%s deviceId=%s status=%s",
            record["id"],
            record["deviceId"],
            record["status"],
        )

    return len(records)


def count_items(
    container,
    where_clause: str | None = None,
    parameters: list[dict[str, Any]] | None = None,
) -> int:
    query = "SELECT VALUE COUNT(1) FROM c"
    if where_clause:
        query += f" WHERE {where_clause}"

    result = list(
        container.query_items(
            query=query,
            parameters=parameters or [],
            enable_cross_partition_query=True,
        )
    )
    return int(result[0]) if result else 0


def find_user_by_email(container, email: str) -> dict[str, Any] | None:
    users = list(
        container.query_items(
            query=(
                "SELECT TOP 1 * FROM c "
                "WHERE c.email = @email"
            ),
            parameters=[{"name": "@email", "value": email}],
            partition_key=email,
        )
    )
    return users[0] if users else None


def find_user_by_id(container, user_id: str) -> dict[str, Any] | None:
    users = list(
        container.query_items(
            query=(
                "SELECT TOP 1 * FROM c "
                "WHERE c.id = @id AND c.doc_type = 'user'"
            ),
            parameters=[{"name": "@id", "value": user_id}],
            enable_cross_partition_query=True,
        )
    )
    return users[0] if users else None


def require_auth(req: func.HttpRequest) -> func.HttpResponse | None:
    _, auth_error = get_auth_claims(req)
    return auth_error


def require_role(
    req: func.HttpRequest,
    allowed_roles: set[str],
) -> tuple[dict[str, Any] | None, func.HttpResponse | None]:
    claims, auth_error = get_auth_claims(req)
    if auth_error:
        return None, auth_error

    role = str(claims.get("role", DEFAULT_USER_ROLE)).lower()
    if role not in allowed_roles:
        return None, error_response("Akses admin diperlukan", 403)

    return claims, None


def get_auth_claims(
    req: func.HttpRequest,
) -> tuple[dict[str, Any] | None, func.HttpResponse | None]:
    auth_header = req.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None, error_response("Sesi login diperlukan", 401)

    token = auth_header.removeprefix("Bearer ").strip()
    try:
        return verify_auth_token(token), None
    except ValueError as exc:
        return None, error_response(str(exc), 401)
    except RuntimeError:
        logging.exception("[Auth] Konfigurasi auth belum tersedia")
        return None, error_response("Konfigurasi auth belum tersedia", 500)
    except Exception:
        logging.exception("[Auth] Token validation error")
        return None, error_response("Token tidak valid", 401)


def create_auth_token(user: dict[str, Any]) -> str:
    issued_at = int(time.time())
    payload = {
        "sub": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user.get("role", "user"),
        "iat": issued_at,
        "exp": issued_at + AUTH_TOKEN_TTL_SECONDS,
    }
    header = {"alg": "HS256", "typ": "JWT"}

    header_part = b64url_encode(json.dumps(header, separators=(",", ":")).encode())
    payload_part = b64url_encode(json.dumps(payload, separators=(",", ":")).encode())
    signing_input = f"{header_part}.{payload_part}"
    signature = sign_token(signing_input)
    return f"{signing_input}.{signature}"


def verify_auth_token(token: str) -> dict[str, Any]:
    parts = token.split(".")
    if len(parts) != 3:
        raise ValueError("Token tidak valid")

    signing_input = f"{parts[0]}.{parts[1]}"
    expected_signature = sign_token(signing_input)
    if not hmac.compare_digest(parts[2], expected_signature):
        raise ValueError("Token tidak valid")

    try:
        payload = json.loads(b64url_decode(parts[1]).decode())
    except (json.JSONDecodeError, UnicodeDecodeError):
        raise ValueError("Token tidak valid")

    if int(payload.get("exp", 0)) < int(time.time()):
        raise ValueError("Sesi login sudah kedaluwarsa")

    required_claims = {"sub", "email", "name"}
    if not required_claims.issubset(payload):
        raise ValueError("Token tidak valid")

    return payload


def sign_token(signing_input: str) -> str:
    digest = hmac.new(
        get_auth_secret().encode(),
        signing_input.encode(),
        hashlib.sha256,
    ).digest()
    return b64url_encode(digest)


def get_auth_secret() -> str:
    secret = _required_env("AUTH_TOKEN_SECRET")
    if len(secret) < 32:
        raise RuntimeError("AUTH_TOKEN_SECRET minimal 32 karakter")
    return secret


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode(),
        salt,
        PBKDF2_ITERATIONS,
    )
    return (
        f"pbkdf2_sha256${PBKDF2_ITERATIONS}$"
        f"{b64url_encode(salt)}${b64url_encode(digest)}"
    )


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_raw, salt_raw, digest_raw = stored_hash.split("$")
        if algorithm != "pbkdf2_sha256":
            return False

        iterations = int(iterations_raw)
        salt = b64url_decode(salt_raw)
        expected_digest = b64url_decode(digest_raw)
        actual_digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode(),
            salt,
            iterations,
        )
        return hmac.compare_digest(actual_digest, expected_digest)
    except Exception:
        return False


def public_user(user: dict[str, Any], include_meta: bool = False) -> dict[str, Any]:
    payload = {
        "id": user["id"],
        "name": user["name"],
        "email": user["email"],
        "role": user.get("role", DEFAULT_USER_ROLE),
    }

    if include_meta:
        for key in ("created_at", "updated_at", "last_login_at"):
            if key in user:
                payload[key] = user[key]

    return payload


def validate_name(raw_name: Any) -> str:
    name = str(raw_name or "").strip()
    if len(name) < 2:
        raise ValueError("Nama minimal 2 karakter")
    if len(name) > 80:
        raise ValueError("Nama maksimal 80 karakter")
    return name


def validate_email(raw_email: Any) -> str:
    email = str(raw_email or "").strip().lower()
    if not EMAIL_PATTERN.match(email):
        raise ValueError("Email tidak valid")
    return email


def validate_password(raw_password: Any) -> str:
    password = str(raw_password or "")
    if len(password) < PASSWORD_MIN_LENGTH:
        raise ValueError(f"Password minimal {PASSWORD_MIN_LENGTH} karakter")
    return password


def validate_role(raw_role: Any) -> str:
    role = str(raw_role or "").strip().lower()
    if role not in USER_ROLES:
        allowed = ", ".join(sorted(USER_ROLES))
        raise ValueError(f"Role tidak valid. Gunakan salah satu: {allowed}")
    return role


def build_azure_ops_summary() -> dict[str, Any]:
    subscription_id = optional_env("AZURE_SUBSCRIPTION_ID")
    resource_group = optional_env("AZURE_RESOURCE_GROUP", "RG-Kelompok11")
    function_name = optional_env("AZURE_FUNCTION_APP_NAME", "func-backend-monitoring-k11")
    storage_name = optional_env("AZURE_STORAGE_ACCOUNT_NAME", "stfuncmonitoringk11")
    cosmos_name = optional_env("AZURE_COSMOS_ACCOUNT_NAME", "cosmos-kelompok11-monitoring")
    vm_name = optional_env("AZURE_VM_NAME", "VM-Web-Kelompok11")

    resources = [
        {
            "key": "function",
            "label": "Azure Functions",
            "resource_id": azure_resource_id(
                subscription_id,
                resource_group,
                "Microsoft.Web",
                "sites",
                function_name,
            ),
            "metricnames": "Requests,Http5xx,AverageResponseTime,CpuTime,MemoryWorkingSet",
            "aggregation": "Total,Average",
        },
        {
            "key": "vm",
            "label": "Developer VM",
            "resource_id": azure_resource_id(
                subscription_id,
                resource_group,
                "Microsoft.Compute",
                "virtualMachines",
                vm_name,
            ),
            "metricnames": "Percentage CPU,Available Memory Bytes",
            "aggregation": "Average",
        },
        {
            "key": "storage",
            "label": "Blob Storage",
            "resource_id": azure_resource_id(
                subscription_id,
                resource_group,
                "Microsoft.Storage",
                "storageAccounts",
                storage_name,
            ),
            "metricnames": "Transactions,UsedCapacity",
            "aggregation": "Total,Average",
        },
        {
            "key": "cosmos",
            "label": "Cosmos DB",
            "resource_id": azure_resource_id(
                subscription_id,
                resource_group,
                "Microsoft.DocumentDB",
                "databaseAccounts",
                cosmos_name,
            ),
            "metricnames": "TotalRequests,ServiceAvailability",
            "aggregation": "Total,Average",
        },
    ]

    if not subscription_id:
        return {
            "configured": False,
            "status": "missing_config",
            "message": "AZURE_SUBSCRIPTION_ID belum dikonfigurasi di backend.",
            "resources": azure_resource_placeholders(resources),
            "totals": empty_azure_totals(),
        }

    credential = DefaultAzureCredential()
    token = credential.get_token("https://management.azure.com/.default").token
    results = []
    for resource in resources:
        results.append(fetch_azure_metric_group(resource, token))

    results = merge_app_insights_summary(results)

    return {
        "configured": True,
        "status": "ready",
        "message": "Metrik Azure Monitor diambil dari Management API.",
        "window": "24h",
        "resources": results,
        "totals": summarize_azure_metrics(results),
    }


def build_cloudflare_ops_summary() -> dict[str, Any]:
    token = optional_env("CLOUDFLARE_API_TOKEN")
    zone_id = optional_env("CLOUDFLARE_ZONE_ID")

    if not token or not zone_id:
        return {
            "configured": False,
            "status": "missing_config",
            "message": "CLOUDFLARE_API_TOKEN atau CLOUDFLARE_ZONE_ID belum dikonfigurasi di backend.",
            "totals": empty_cloudflare_totals(),
            "daily": [],
        }

    end_date = datetime.now(timezone.utc).date()
    start_date = end_date - timedelta(days=6)
    query = """
    query ZoneTraffic($zoneTag: string, $start: Date, $end: Date) {
      viewer {
        zones(filter: { zoneTag: $zoneTag }) {
          httpRequests1dGroups(
            limit: 7,
            filter: { date_geq: $start, date_leq: $end },
            orderBy: [date_ASC]
          ) {
            dimensions { date }
            sum { requests pageViews bytes cachedBytes threats }
            uniq { uniques }
          }
        }
      }
    }
    """
    payload = {
        "query": query,
        "variables": {
            "zoneTag": zone_id,
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        },
    }

    try:
        response = http_json(
            "https://api.cloudflare.com/client/v4/graphql",
            payload=payload,
            headers={"Authorization": f"Bearer {token}"},
        )
    except Exception as exc:
        logging.warning("[Admin Ops] Gagal mengambil Cloudflare analytics: %s", exc)
        return {
            "configured": True,
            "status": "error",
            "message": "Cloudflare Analytics belum dapat diakses oleh backend.",
            "detail": str(exc)[:180],
            "totals": empty_cloudflare_totals(),
            "daily": [],
        }

    if response.get("errors"):
        return {
            "configured": True,
            "status": "error",
            "message": "Cloudflare GraphQL mengembalikan error.",
            "errors": response.get("errors", [])[:2],
            "totals": empty_cloudflare_totals(),
            "daily": [],
        }

    groups = (
        response.get("data", {})
        .get("viewer", {})
        .get("zones", [{}])[0]
        .get("httpRequests1dGroups", [])
    )
    daily = [normalize_cloudflare_day(item) for item in groups]
    totals = {
        "requests": sum(item["requests"] for item in daily),
        "page_views": sum(item["page_views"] for item in daily),
        "unique_visitors": sum(item["unique_visitors"] for item in daily),
        "bandwidth_bytes": sum(item["bandwidth_bytes"] for item in daily),
        "cached_bytes": sum(item["cached_bytes"] for item in daily),
        "threats": sum(item["threats"] for item in daily),
    }
    totals["cache_ratio"] = round(
        (totals["cached_bytes"] / totals["bandwidth_bytes"]) * 100,
        2,
    ) if totals["bandwidth_bytes"] else 0

    return {
        "configured": True,
        "status": "ready",
        "message": "Traffic web diambil dari Cloudflare GraphQL Analytics.",
        "window": "7d",
        "totals": totals,
        "daily": daily,
    }


def azure_resource_id(
    subscription_id: str,
    resource_group: str,
    namespace: str,
    resource_type: str,
    resource_name: str,
) -> str:
    return (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/{namespace}/{resource_type}/{resource_name}"
    )


def fetch_azure_metric_group(resource: dict[str, Any], token: str) -> dict[str, Any]:
    try:
        response = http_json(
            (
                "https://management.azure.com"
                f"{resource['resource_id']}/providers/microsoft.insights/metrics"
            ),
            params={
                "api-version": "2018-01-01",
                "metricnames": resource["metricnames"],
                "timespan": azure_timespan(hours=24),
                "interval": "PT1H",
                "aggregation": resource["aggregation"],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        metrics = {
            metric.get("name", {}).get("value"): summarize_azure_metric(metric)
            for metric in response.get("value", [])
        }
        return {
            "key": resource["key"],
            "label": resource["label"],
            "status": "ready",
            "metrics": metrics,
        }
    except Exception as exc:
        logging.warning("[Admin Ops] Gagal mengambil metrik %s: %s", resource["key"], exc)
        return {
            "key": resource["key"],
            "label": resource["label"],
            "status": "error",
            "message": str(exc),
            "metrics": {},
        }


def summarize_azure_metric(metric: dict[str, Any]) -> dict[str, Any]:
    points = []
    for series in metric.get("timeseries", []):
        for item in series.get("data", []):
            value = first_number(item, ("total", "average", "maximum", "minimum", "count"))
            if value is not None:
                points.append(
                    {
                        "time": item.get("timeStamp"),
                        "value": value,
                    }
                )

    values = [point["value"] for point in points]
    unit = metric.get("unit", "Count")
    return {
        "display_name": metric.get("displayDescription") or metric.get("name", {}).get("localizedValue"),
        "unit": unit,
        "latest": values[-1] if values else 0,
        "total": round(sum(values), 3) if values else 0,
        "average": round(sum(values) / len(values), 3) if values else 0,
        "points": points[-24:],
    }


def summarize_azure_metrics(resources: list[dict[str, Any]]) -> dict[str, Any]:
    metrics = {}
    for resource in resources:
        metrics.update(resource.get("metrics", {}))

    requests = metrics.get("Requests", {}).get("total", 0)
    http_5xx = metrics.get("Http5xx", {}).get("total", 0)
    latency = metrics.get("AverageResponseTime", {}).get("average", 0)
    cpu_percent = metrics.get("Percentage CPU", {}).get("average", 0)
    cpu_time = metrics.get("CpuTime", {}).get("total", 0)
    memory_working_set = first_metric_value(metrics.get("MemoryWorkingSet", {}))
    available_memory = first_metric_value(metrics.get("Available Memory Bytes", {}))
    storage_transactions = metrics.get("Transactions", {}).get("total", 0)
    cosmos_requests = metrics.get("TotalRequests", {}).get("total", 0)
    availability = metrics.get("ServiceAvailability", {}).get("average", 0)
    request_rate = round(requests / (24 * 60), 3) if requests else 0
    error_rate = round((http_5xx / requests) * 100, 3) if requests else 0

    return {
        "function_requests": requests,
        "function_5xx": http_5xx,
        "function_avg_response_time": latency,
        "function_request_rate_per_minute": request_rate,
        "function_error_rate": error_rate,
        "cpu_usage_percent": cpu_percent,
        "cpu_time_seconds": cpu_time,
        "memory_working_set_bytes": memory_working_set,
        "available_memory_bytes": available_memory,
        "storage_transactions": storage_transactions,
        "cosmos_requests": cosmos_requests,
        "cosmos_availability": availability,
    }


def merge_app_insights_summary(resources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    app_id = optional_env("APPINSIGHTS_APP_ID")
    api_key = optional_env("APPINSIGHTS_API_KEY")
    if not app_id or not api_key:
        return resources

    try:
        summary = fetch_app_insights_summary(app_id, api_key)
    except Exception as exc:
        logging.warning("[Admin Ops] Gagal mengambil Application Insights summary: %s", exc)
        return resources

    merged = list(resources)
    function_resource = next(
        (resource for resource in merged if resource.get("key") == "function"),
        None,
    )
    if not function_resource:
        return merged

    function_resource["status"] = "ready"
    function_resource["message"] = "Fallback Application Insights API"
    metrics = function_resource.setdefault("metrics", {})
    metrics["Requests"] = {
        "display_name": "Requests",
        "unit": "Count",
        "latest": summary["requests_points"][-1]["value"] if summary["requests_points"] else 0,
        "total": summary["requests"],
        "average": summary["request_rate_per_hour"],
        "points": summary["requests_points"],
    }
    metrics["Http5xx"] = {
        "display_name": "HTTP 5xx",
        "unit": "Count",
        "latest": summary["error_points"][-1]["value"] if summary["error_points"] else 0,
        "total": summary["errors"],
        "average": summary["errors"] / 24 if summary["errors"] else 0,
        "points": summary["error_points"],
    }
    metrics["AverageResponseTime"] = {
        "display_name": "Average response time",
        "unit": "Seconds",
        "latest": summary["avg_duration_seconds"],
        "total": summary["avg_duration_seconds"],
        "average": summary["avg_duration_seconds"],
        "points": [
            {
                "time": datetime.now(timezone.utc).isoformat(),
                "value": summary["avg_duration_seconds"],
            }
        ],
    }
    metrics["Percentage CPU"] = {
        "display_name": "Process CPU",
        "unit": "Percent",
        "latest": summary["cpu_percent"],
        "total": summary["cpu_percent"],
        "average": summary["cpu_percent"],
        "points": [
            {
                "time": datetime.now(timezone.utc).isoformat(),
                "value": summary["cpu_percent"],
            }
        ],
    }
    metrics["MemoryWorkingSet"] = {
        "display_name": "Private bytes",
        "unit": "Bytes",
        "latest": summary["memory_private_bytes"],
        "total": summary["memory_private_bytes"],
        "average": summary["memory_private_bytes"],
        "points": [
            {
                "time": datetime.now(timezone.utc).isoformat(),
                "value": summary["memory_private_bytes"],
            }
        ],
    }
    return merged


def fetch_app_insights_summary(app_id: str, api_key: str) -> dict[str, Any]:
    query = """
    requests
    | where timestamp > ago(24h)
    | summarize requests=count(), errors=countif(tolong(resultCode) >= 500), avgDurationMs=avg(duration)
    """
    summary = app_insights_query(app_id, api_key, query)
    row = first_app_insights_row(summary)
    requests = int(first_number_from_row(row, 0) or 0)
    errors = int(first_number_from_row(row, 1) or 0)
    avg_duration_ms = first_number_from_row(row, 2) or 0

    series_query = """
    requests
    | where timestamp > ago(24h)
    | summarize requests=count(), errors=countif(tolong(resultCode) >= 500) by bin(timestamp, 1h)
    | order by timestamp asc
    """
    series = app_insights_query(app_id, api_key, series_query)
    requests_points = []
    error_points = []
    for series_row in app_insights_rows(series):
        timestamp = series_row[0] if len(series_row) > 0 else None
        requests_value = first_number_from_row(series_row, 1) or 0
        errors_value = first_number_from_row(series_row, 2) or 0
        requests_points.append({"time": timestamp, "value": requests_value})
        error_points.append({"time": timestamp, "value": errors_value})

    performance_query = """
    performanceCounters
    | where timestamp > ago(24h)
    | where category == "Process" and counter in ("% Processor Time Normalized", "Private Bytes")
    | summarize avgValue=avg(value) by counter
    """
    performance = app_insights_query(app_id, api_key, performance_query)
    cpu_percent = 0
    memory_private_bytes = 0
    for performance_row in app_insights_rows(performance):
        counter = str(performance_row[0] if performance_row else "")
        value = first_number_from_row(performance_row, 1) or 0
        if counter == "% Processor Time Normalized":
            cpu_percent = round(value, 3)
        elif counter == "Private Bytes":
            memory_private_bytes = round(value, 3)

    return {
        "requests": requests,
        "errors": errors,
        "avg_duration_seconds": round(avg_duration_ms / 1000, 3) if avg_duration_ms else 0,
        "request_rate_per_hour": round(requests / 24, 3) if requests else 0,
        "requests_points": requests_points[-24:],
        "error_points": error_points[-24:],
        "cpu_percent": cpu_percent,
        "memory_private_bytes": memory_private_bytes,
    }


def app_insights_query(app_id: str, api_key: str, query: str) -> dict[str, Any]:
    return http_json(
        f"https://api.applicationinsights.io/v1/apps/{app_id}/query",
        payload={"query": query},
        headers={"x-api-key": api_key},
    )


def first_app_insights_row(payload: dict[str, Any]) -> list[Any]:
    rows = app_insights_rows(payload)
    return rows[0] if rows else []


def app_insights_rows(payload: dict[str, Any]) -> list[list[Any]]:
    tables = payload.get("tables") or []
    if not tables:
        return []
    rows = tables[0].get("rows") or []
    return rows if isinstance(rows, list) else []


def first_number_from_row(row: list[Any], index: int) -> float | None:
    if len(row) <= index:
        return None
    value = row[index]
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    return None


def normalize_cloudflare_day(item: dict[str, Any]) -> dict[str, Any]:
    summary = item.get("sum", {})
    uniq = item.get("uniq", {})
    return {
        "date": item.get("dimensions", {}).get("date"),
        "requests": int(summary.get("requests") or 0),
        "page_views": int(summary.get("pageViews") or 0),
        "unique_visitors": int(uniq.get("uniques") or 0),
        "bandwidth_bytes": int(summary.get("bytes") or 0),
        "cached_bytes": int(summary.get("cachedBytes") or 0),
        "threats": int(summary.get("threats") or 0),
    }


def azure_timespan(hours: int) -> str:
    end = datetime.now(timezone.utc)
    start = end - timedelta(hours=hours)
    return f"{start.isoformat()}/{end.isoformat()}"


def http_json(
    url: str,
    payload: dict[str, Any] | None = None,
    params: dict[str, str] | None = None,
    headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    if params:
        query = "&".join(
            f"{url_quote(str(key))}={url_quote(str(value))}"
            for key, value in params.items()
        )
        url = f"{url}?{query}"

    body = None
    request_headers = {"Accept": "application/json", **(headers or {})}
    if payload is not None:
        body = json.dumps(payload).encode()
        request_headers["Content-Type"] = "application/json"

    request = Request(url, data=body, headers=request_headers, method="POST" if body else "GET")
    try:
        with urlopen(request, timeout=12) as response:
            return json.loads(response.read().decode())
    except HTTPError as exc:
        detail = exc.read().decode(errors="replace")[:500]
        raise RuntimeError(f"HTTP {exc.code}: {detail}") from exc
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc


def url_quote(value: str) -> str:
    safe = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_.~:/,"
    return "".join(char if char in safe else f"%{ord(char):02X}" for char in value)


def first_number(item: dict[str, Any], keys: tuple[str, ...]) -> float | None:
    for key in keys:
        value = item.get(key)
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            return float(value)
    return None


def first_metric_value(metric: dict[str, Any]) -> float:
    for key in ("latest", "average", "total"):
        value = metric.get(key)
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            return float(value)
    return 0


def azure_resource_placeholders(resources: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "key": resource["key"],
            "label": resource["label"],
            "status": "missing_config",
            "metrics": {},
        }
        for resource in resources
    ]


def empty_azure_totals() -> dict[str, int | float]:
    return {
        "function_requests": 0,
        "function_5xx": 0,
        "function_avg_response_time": 0,
        "function_request_rate_per_minute": 0,
        "function_error_rate": 0,
        "cpu_usage_percent": 0,
        "cpu_time_seconds": 0,
        "memory_working_set_bytes": 0,
        "available_memory_bytes": 0,
        "storage_transactions": 0,
        "cosmos_requests": 0,
        "cosmos_availability": 0,
    }


def empty_cloudflare_totals() -> dict[str, int | float]:
    return {
        "requests": 0,
        "page_views": 0,
        "unique_visitors": 0,
        "bandwidth_bytes": 0,
        "cached_bytes": 0,
        "cache_ratio": 0,
        "threats": 0,
    }


def b64url_encode(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode().rstrip("=")


def b64url_decode(value: str) -> bytes:
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode(f"{value}{padding}".encode())


def parse_limit(raw_limit: str | None) -> int:
    if raw_limit is None:
        return DEFAULT_LIMIT

    try:
        limit = int(raw_limit)
    except (TypeError, ValueError):
        raise ValueError("Parameter limit harus berupa angka")

    if limit < 1:
        raise ValueError("Parameter limit minimal 1")
    if limit > MAX_LIMIT:
        raise ValueError(f"Parameter limit maksimal {MAX_LIMIT}")

    return limit


def parse_status(raw_status: str | None) -> str | None:
    if raw_status is None or raw_status == "":
        return None

    status = raw_status.lower()
    if status not in ALLOWED_STATUS:
        allowed = ", ".join(sorted(ALLOWED_STATUS))
        raise ValueError(f"Status tidak valid. Gunakan salah satu: {allowed}")

    return status


def is_valid_payload(data: Any) -> bool:
    return data is not None and isinstance(data, (dict, list))


def extract_device_id(item: Any) -> str:
    if isinstance(item, dict):
        for key in ("deviceId", "device_id", "deviceid", "device", "host", "source"):
            value = item.get(key)
            if value:
                return str(value)

    return "unknown-device"


def optional_env(name: str, default: str = "") -> str:
    return os.environ.get(name, default).strip()


def _required_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Environment variable {name} belum dikonfigurasi")
    return value
